import io
import os
import re
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import Workbook

# --- 1. CẤU HÌNH STREAMLIT ---
st.set_page_config(page_title="Tool nhập liệu DVCTECH", layout="wide")

# --- 2. CSS TÙY CHỈNH GIAO DIỆN ---
st.markdown(
    """
    <style>
    div.stDownloadButton > button {
        background-color: DarkGreen !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 0.5rem 1.2rem !important;
    }
    div.stDownloadButton > button:hover {
        background-color: Green !important;
        color: #FFFFFF !important;
    }
    div.stDownloadButton > button p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    div.stDownloadButton > button[kind="primary"] {
        background-color: DarkRed !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 0.5rem 1.2rem !important;
    }
    div.stDownloadButton > button[kind="primary"]:hover {
        background-color: Red !important;
        color: #FFFFFF !important;
    }
    div.stDownloadButton > button[kind="primary"] p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    div[data-testid="stFileUploader"] button {
        background-color: darkblue !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border-radius: 6px !important;
        border: none !important;
    }
    div[data-testid="stFileUploader"] button:hover {
        background-color: #0066CC !important;
        color: #FFFFFF !important;
    }
    div[data-testid="stFileUploader"] p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    div[data-testid="stMarkdownContainer"] p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    div.stButton > button[kind="primary"] {
        background-color: DarkBlue  !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 0.5rem 1.2rem !important;
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #0066CC !important;
        color: #FFFFFF !important;
    }
    div.stButton > button[kind="primary"] p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    div.stButton > button[kind="secondary"] {
        background-color: Gray !important;
        color: #FFFFFF !important;
        font-weight: bold !important;
        border-radius: 6px !important;
        border: none !important;
        padding: 0.5rem 1.2rem !important;
    }
    div.stButton > button[kind="secondary"]:hover {
        background-color: DarkGray !important;
        color: #FFFFFF !important;
    }
    div.stButton > button[kind="secondary"] p {
        font-size: 18px !important;
        font-weight: bold !important;
    }
    </style>
""",
    unsafe_allow_html=True,
)

st.title("TOOL NHẬP LIỆU BẢNG BÁO GIÁ")


# --- 3. HÀM TẠO FILE FORM MẪU ---
def generate_sample_template():
    sample_df = pd.DataFrame({
        "Stt": [],
        "Thiết bị": [],
        "Mã hàng": [],
        "Hãng/\nXuất xứ": [],
        "Mô tả": [],
        "ĐVT": [],
        "Số lượng": [],
        "Thời gian bảo hành": [],
        "Ghi chú": [],
        "Margin Thiết bị": [],
        "ĐG COST Thiết bị": [],
        "Margin Lắp đặt": [],
        "ĐG COST Lắp đặt": [],
        "NCC": [],
        "NOTE": [],
    })

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sample_df.to_excel(writer, index=False, sheet_name="FORM_MAU")

        worksheet = writer.sheets["FORM_MAU"]
        gray_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col in range(1, 16):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = gray_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in range(2, 4):
            for col in range(1, 16):
                worksheet.cell(row=row, column=col).border = thin_border

    return output.getvalue()


# --- 4. HÀM LÀM SẠCH DỮ LIỆU SỐ & CÔNG THỨC ---
def clean_currency(val):
    if pd.isna(val) or val == "" or val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)

    val_str = str(val).strip()
    if "%" in val_str:
        val_str = val_str.replace("%", "").strip()
        try:
            return float(val_str) / 100.0
        except ValueError:
            return 0.0

    clean_str = re.sub(r"[^\d.]", "", val_str.replace(",", "."))
    try:
        return float(clean_str) if clean_str else 0.0
    except ValueError:
        return 0.0


def parse_formula_or_value(val):
    """Xử lý giá trị nhập vào cho các cột Cost:
    Nếu nhập công thức toán học (vd: 100000/1.08 hoặc =100000/1.08), giữ nguyên dạng công thức Excel.
    Nếu nhập số thông thường, trả về dạng số thực.
    """
    if pd.isna(val) or val == "" or val is None:
        return 0
    if isinstance(val, (int, float)):
        return val

    val_str = str(val).strip()
    if not val_str:
        return 0

    # Nếu bắt đầu bằng dấu '='
    if val_str.startswith("="):
        return val_str.replace(",", ".")

    # Nếu chứa các phép tính +, -, *, /
    if any(op in val_str for op in ["/", "*", "+", "-"]):
        clean_formula = val_str.replace(",", ".")
        return f"={clean_formula}"

    # Nếu là số bình thường
    return clean_currency(val_str)


def parse_margin(val):
    """Xử lý giá trị Margin nhập vào (hỗ trợ nhập 10, 10% hoặc 0.1)."""
    clean_val = clean_currency(val)
    if clean_val >= 1.0:
        return clean_val / 100.0
    return clean_val


# --- 5. HÀM Chuẩn hóa 1 Dataframe đơn lẻ ---
def standardize_df(input_df):
    def get_col_val(df, possible_names, default=""):
        for name in possible_names:
            for c in df.columns:
                c_clean = c.lower().replace("/", "").replace(" ", "")
                target_clean = name.lower().replace("/", "").replace(" ", "")
                if target_clean in c_clean:
                    return df[c]
        return pd.Series([default] * len(df))

    df_final = pd.DataFrame()
    df_final["STT"] = get_col_val(input_df, ["stt"], 1)
    df_final["Thiết bị"] = get_col_val(
        input_df,
        ["thiết bị", "tên thiết bị", "tên hàng hóa/dịch vụ", "tên hàng hóa"],
        "",
    )
    df_final["Mã hàng"] = get_col_val(input_df, ["mã hàng"], "")
    df_final["Hãng/\nXuất xứ"] = get_col_val(
        input_df, ["hãng/xuất xứ", "nhãn hiệu/xuất xứ", "xuất xứ", "hãng"], ""
    )
    df_final["Mô tả"] = get_col_val(input_df, ["mô tả"], "")
    df_final["Hình ảnh"] = ""
    df_final["ĐVT"] = get_col_val(input_df, ["đvt"], "Cái")

    raw_sl = get_col_val(input_df, ["số lượng"], 1)
    df_final["Số lượng"] = raw_sl.apply(clean_currency)

    df_final["Đơn giá (VNĐ)"] = None
    df_final["Thành tiền (VNĐ)"] = None
    df_final["Thời gian bảo hành"] = get_col_val(
        input_df, ["thời gian bảo hành", "bảo hành"], ""
    )
    df_final["Ghi chú"] = get_col_val(input_df, ["ghi chú"], "")

    raw_margin = get_col_val(input_df, ["margin thiết bị", "margin tb"], 0)
    df_final["Margin Thiết bị"] = raw_margin.apply(parse_margin)

    raw_cost = get_col_val(
        input_df, ["đg cost thiết bị", "cost thiết bị", "giá cost"], 0
    )
    df_final["ĐG COST Thiết bị"] = raw_cost.apply(parse_formula_or_value)
    df_final["TT COST Thiết bị"] = None

    raw_margin_ld = get_col_val(input_df, ["margin lắp đặt", "margin ld"], 0)
    df_final["Margin Lắp đặt"] = raw_margin_ld.apply(parse_margin)

    raw_lapdat = get_col_val(
        input_df, ["đg cost lắp đặt", "cost lắp đặt", "giá lắp đặt"], 0
    )
    df_final["ĐG COST Lắp đặt"] = raw_lapdat.apply(parse_formula_or_value)
    df_final["TT COST Lắp đặt"] = None

    df_final["NCC"] = get_col_val(input_df, ["ncc"], "")
    df_final["NOTE"] = get_col_val(input_df, ["note"], "")

    form_columns = [
        "STT", "Thiết bị", "Mã hàng", "Hãng/\nXuất xứ", "Mô tả", "ĐVT",
        "Số lượng", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)", "Thời gian bảo hành",
        "Ghi chú", "Margin Thiết bị", "ĐG COST Thiết bị", "TT COST Thiết bị",
        "Margin Lắp đặt", "ĐG COST Lắp đặt", "TT COST Lắp đặt", "NCC", "NOTE"
    ]
    return df_final.reindex(columns=form_columns)


# --- 6. HÀM XỬ LÝ DỮ LIỆU VÀ TẠO FILE EXCEL HOÀN CHỈNH (3 SHEETS) ---
def process_dataframe_and_generate_excel(raw_input_df):
    cols = [str(c).replace("\n", " ").strip() for c in raw_input_df.columns]
    raw_input_df.columns = cols

    split_idx = None
    for idx, row in raw_input_df.iterrows():
        device_val = str(row.get("Thiết bị", "")).strip()
        if device_val == "" or pd.isna(row.get("Thiết bị")):
            remaining = raw_input_df.iloc[idx + 1 :]
            if not remaining.empty and remaining["Thiết bị"].dropna().astype(str).str.strip().ne("").any():
                split_idx = idx
                break

    if split_idx is not None:
        raw_sec1 = raw_input_df.iloc[:split_idx].copy()
        raw_sec2 = raw_input_df.iloc[split_idx + 1 :].copy()
    else:
        raw_sec1 = raw_input_df.copy()
        raw_sec2 = pd.DataFrame()

    if "Thiết bị" in raw_sec1.columns:
        raw_sec1 = raw_sec1[raw_sec1["Thiết bị"].dropna().astype(str).str.strip().ne("")].reset_index(drop=True)
    if "Thiết bị" in raw_sec2.columns and not raw_sec2.empty:
        raw_sec2 = raw_sec2[raw_sec2["Thiết bị"].dropna().astype(str).str.strip().ne("")].reset_index(drop=True)

    df_sec1 = standardize_df(raw_sec1)
    df_sec2 = standardize_df(raw_sec2)

    if not df_sec1.empty:
        df_sec1["STT"] = range(1, len(df_sec1) + 1)
    if not df_sec2.empty:
        df_sec2["STT"] = range(1, len(df_sec2) + 1)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        ws_bg = writer.book.create_sheet(title="BÁO GIÁ", index=0)
        ws_ct = writer.book.create_sheet(title="CHI TIẾT", index=1)

        form_columns = [
            "STT", "Thiết bị", "Mã hàng", "Hãng/\nXuất xứ", "Mô tả", "ĐVT",
            "Số lượng", "Đơn giá (VNĐ)", "Thành tiền (VNĐ)", "Thời gian\nbảo hành",
            "Ghi chú", "Margin\nThiết bị", "ĐG COST\nThiết bị", "TT COST\nThiết bị",
            "Margin\nLắp đặt", "ĐG COST\nLắp đặt", "TT COST\nLắp đặt", "NCC", "NOTE"
        ]

        for col_idx, col_name in enumerate(form_columns, 1):
            ws_ct.cell(row=4, column=col_idx, value=col_name)

        ws_ct.merge_cells("B2:J2")
        title_ct = ws_ct["B2"]
        title_ct.value = "BẢNG GIÁ CHI TIẾT"
        title_ct.font = Font(name="Times New Roman", size=18, bold=True)
        title_ct.alignment = Alignment(horizontal="center", vertical="center")

        num_format_vnd = "#,##0"

        # --- GHI MỤC I ---
        ws_ct.cell(row=5, column=1, value="I").alignment = Alignment(horizontal="center", vertical="center")
        ws_ct.cell(row=5, column=2, value="Hàng hóa/Thiết bị chính").font = Font(name="Times New Roman", size=10, bold=True)

        n_sec1 = len(df_sec1)
        start_r_sec1 = 6
        end_r_sec1 = start_r_sec1 + n_sec1 - 1 if n_sec1 > 0 else start_r_sec1

        if n_sec1 > 0:
            for i, (_, row_data) in enumerate(df_sec1.iterrows()):
                r = start_r_sec1 + i
                for c_idx, val in enumerate(row_data, 1):
                    ws_ct.cell(row=r, column=c_idx, value=val)
            ws_ct.cell(row=5, column=9, value=f"=SUM(I{start_r_sec1}:I{end_r_sec1})").number_format = num_format_vnd
        else:
            ws_ct.cell(row=5, column=9, value=0).number_format = num_format_vnd

        ws_ct.cell(row=5, column=9).font = Font(name="Times New Roman", size=10, bold=True)

        for r in range(start_r_sec1, start_r_sec1 + n_sec1):
            ws_ct.cell(row=r, column=8, value=f"=ROUNDUP(M{r}/(1-L{r}), -3)").number_format = num_format_vnd
            ws_ct.cell(row=r, column=9, value=f"=G{r}*H{r}").number_format = num_format_vnd
            ws_ct.cell(row=r, column=14, value=f"=G{r}*M{r}").number_format = num_format_vnd
            ws_ct.cell(row=r, column=17, value=f"=G{r}*P{r}").number_format = num_format_vnd

        # --- GHI MỤC II ---
        row_II = end_r_sec1 + 1 if n_sec1 > 0 else 6
        ws_ct.cell(row=row_II, column=1, value="II").alignment = Alignment(horizontal="center", vertical="center")
        ws_ct.cell(row=row_II, column=2, value="Chi phí triển khai").font = Font(name="Times New Roman", size=10, bold=True)

        n_sec2 = len(df_sec2)
        start_r_sec2 = row_II + 1
        end_r_sec2 = start_r_sec2 + n_sec2 - 1 if n_sec2 > 0 else start_r_sec2

        if n_sec2 > 0:
            for i, (_, row_data) in enumerate(df_sec2.iterrows()):
                r = start_r_sec2 + i
                for c_idx, val in enumerate(row_data, 1):
                    ws_ct.cell(row=r, column=c_idx, value=val)
            ws_ct.cell(row=row_II, column=9, value=f"=SUM(I{start_r_sec2}:I{end_r_sec2})").number_format = num_format_vnd

            for r in range(start_r_sec2, start_r_sec2 + n_sec2):
                ws_ct.cell(row=r, column=8, value=f"=ROUNDUP(M{r}/(1-L{r}), -3)").number_format = num_format_vnd
                ws_ct.cell(row=r, column=9, value=f"=G{r}*H{r}").number_format = num_format_vnd
                ws_ct.cell(row=r, column=14, value=f"=G{r}*M{r}").number_format = num_format_vnd
                ws_ct.cell(row=r, column=17, value=f"=G{r}*P{r}").number_format = num_format_vnd
        else:
            ws_ct.cell(row=row_II, column=6, value="Gói").alignment = Alignment(horizontal="center", vertical="center")
            ws_ct.cell(row=row_II, column=7, value=1).alignment = Alignment(horizontal="center", vertical="center")
            ws_ct.cell(row=row_II, column=8, value=0).number_format = num_format_vnd
            ws_ct.cell(row=row_II, column=9, value=f"=G{row_II}*H{row_II}").number_format = num_format_vnd

        ws_ct.cell(row=row_II, column=9).font = Font(name="Times New Roman", size=10, bold=True)

        # --- DÒNG TỔNG CỘNG ---
        tot_row_ct = end_r_sec2 + 1 if n_sec2 > 0 else row_II + 1
        ws_ct.merge_cells(start_row=tot_row_ct, start_column=1, end_row=tot_row_ct, end_column=8)
        cell_tot = ws_ct.cell(row=tot_row_ct, column=1, value="TỔNG CỘNG")
        cell_tot.font = Font(name="Times New Roman", size=10, bold=True)
        cell_tot.alignment = Alignment(horizontal="center", vertical="center")

        ws_ct.cell(row=tot_row_ct, column=9, value=f"=I5+I{row_II}").number_format = num_format_vnd
        ws_ct.cell(row=tot_row_ct, column=9).font = Font(name="Times New Roman", size=10, bold=True)

        col_widths_ct = {
            "A": 5, "B": 28, "C": 10, "D": 10, "E": 18,
            "F": 5, "G": 9, "H": 10, "I": 12, "J": 12, "K": 12,
            "L": 8, "M": 10, "N": 10, "O": 8, "P": 10, "Q": 10, "R": 7, "S": 8
        }
        for col_letter, width in col_widths_ct.items():
            ws_ct.column_dimensions[col_letter].width = width

        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for col in range(1, 12):
            cell = ws_ct.cell(row=4, column=col)
            cell.fill, cell.border = gray_fill, thin_border
            cell.font = Font(name="Times New Roman", size=10, bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(12, 20):
            cell = ws_ct.cell(row=4, column=col)
            cell.fill, cell.border = gray_fill, thin_border
            cell.font = Font(name="Times New Roman", size=10, bold=True, color="FF0000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_right = Alignment(horizontal="right", vertical="center")
        num_format_percent = "0%"

        for r in range(5, tot_row_ct + 1):
            is_header = r in (5, row_II, tot_row_ct)
            for c in range(1, 20):
                cell = ws_ct.cell(row=r, column=c)
                cell.border = thin_border
                cell.font = Font(name="Times New Roman", size=10, bold=is_header)

                if c in [1, 3, 4, 6, 7, 10, 12, 15, 18]:
                    cell.alignment = align_center
                elif c in [2, 5, 11, 19]:
                    cell.alignment = align_left
                else:
                    cell.alignment = align_right

                if c in [12, 15]:
                    cell.number_format = num_format_percent
                elif c in [8, 9, 13, 14, 16, 17]:
                    cell.number_format = num_format_vnd

                if r == tot_row_ct:
                    cell.fill = gray_fill

        # =========================================================
        # B. TẠO VÀ XỬ LÝ SHEET GUI_KH
        # =========================================================
        ws_kh = writer.book.create_sheet(title="GUI_KH", index=2)
        ws_kh.merge_cells("A2:K2")
        ws_kh["A2"] = "BẢNG GIÁ CHI TIẾT"
        ws_kh["A2"].font = Font(name="Times New Roman", size=18, bold=True)
        ws_kh["A2"].alignment = Alignment(horizontal="center", vertical="center")

        headers_kh = [
            ("A4", "STT"), ("B4", "Thiết bị"), ("C4", "Mã hàng"), 
            ("D4", "Hãng/\nXuất xứ"), ("E4", "Mô tả"), ("F4", "ĐVT"), 
            ("G4", "Số lượng"), ("H4", "Đơn giá\n(VNĐ)"), 
            ("I4", "Thành tiền\n(VNĐ)"), ("J4", "Thời gian\nbảo hành"), ("K4", "Ghi chú")
        ]
        ws_kh.row_dimensions[4].height = 28
        for cell_id, text in headers_kh:
            c = ws_kh[cell_id]
            c.value = text
            c.font = Font(name="Times New Roman", size=10, bold=True)
            c.fill = gray_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border

        for r in range(5, tot_row_ct):
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "J", "K"]:
                ws_kh[f"{col_letter}{r}"] = f"=IF('CHI TIẾT'!{col_letter}{r}=\"\",\"\",'CHI TIẾT'!{col_letter}{r})"

            if r == 5:
                ws_kh[f"I{r}"] = f"=SUM(I{start_r_sec1}:I{end_r_sec1})" if n_sec1 > 0 else 0
            elif r == row_II:
                ws_kh[f"I{r}"] = f"=SUM(I{start_r_sec2}:I{end_r_sec2})" if n_sec2 > 0 else f"=G{r}*H{r}"
            else:
                ws_kh[f"I{r}"] = f"=G{r}*H{r}"

            is_bold = r in (5, row_II)
            for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
                c = ws_kh[f"{col_letter}{r}"]
                c.font = Font(name="Times New Roman", size=10, bold=is_bold)
                c.border = thin_border
                if col_letter in ["A", "C", "D", "F", "G", "J"]:
                    c.alignment = align_center
                elif col_letter in ["H", "I"]:
                    c.alignment = align_right
                    c.number_format = num_format_vnd
                else:
                    c.alignment = align_left

        ws_kh.merge_cells(f"A{tot_row_ct}:H{tot_row_ct}")
        ws_kh[f"A{tot_row_ct}"] = "TỔNG CỘNG"
        ws_kh[f"A{tot_row_ct}"].font = Font(name="Times New Roman", size=10, bold=True)
        ws_kh[f"A{tot_row_ct}"].alignment = align_center

        ws_kh[f"I{tot_row_ct}"] = f"=I5+I{row_II}"
        ws_kh[f"I{tot_row_ct}"].font = Font(name="Times New Roman", size=10, bold=True)
        ws_kh[f"I{tot_row_ct}"].alignment = align_right
        ws_kh[f"I{tot_row_ct}"].number_format = num_format_vnd

        for col_idx in range(1, 12):
            col_letter = get_column_letter(col_idx)
            cell = ws_kh[f"{col_letter}{tot_row_ct}"]
            cell.border = thin_border
            cell.fill = gray_fill

        col_widths_kh = {
            "A": 5, "B": 28, "C": 10, "D": 10, "E": 18, 
            "F": 5, "G": 9, "H": 10, "I": 12, "J": 12, "K": 12
        }
        for col_letter, width in col_widths_kh.items():
            ws_kh.column_dimensions[col_letter].width = width

        # =========================================================
        # C. TẠO VÀ XỬ LÝ SHEET BÁO GIÁ
        # =========================================================
        ws_bg.page_setup.orientation = ws_bg.ORIENTATION_PORTRAIT
        ws_bg.page_setup.paperSize = ws_bg.PAPERSIZE_A4
        ws_bg.sheet_properties.pageSetUpPr.fitToPage = True
        ws_bg.page_setup.fitToWidth = 1
        ws_bg.page_setup.fitToHeight = 0

        try:
            img = Image("logo_dvc.png")
            img.width = 90
            img.height = 90
            ws_bg.add_image(img, "A1")
        except Exception:
            pass

        ws_bg.merge_cells("A1:G1")
        ws_bg["A1"] = "CÔNG TY TNHH CÔNG NGHỆ DVC"
        ws_bg["A1"].font = Font(name="Times New Roman", size=12, bold=True)
        ws_bg["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg.merge_cells("A2:G2")
        ws_bg["A2"] = "**********"
        ws_bg["A2"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg.merge_cells("A3:G3")
        ws_bg["A3"] = "Hotline: 0909 661 579"
        ws_bg["A3"].font = Font(name="Times New Roman", size=10)
        ws_bg["A3"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg.merge_cells("A4:G4")
        ws_bg["A4"] = "Email: dvc@dvctech.vn - Website: dvctech.vn"
        ws_bg["A4"].font = Font(name="Times New Roman", size=10, underline="single")
        ws_bg["A4"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg.merge_cells("A5:G5")
        ws_bg["A5"] = "BẢNG BÁO GIÁ"
        ws_bg["A5"].font = Font(name="Times New Roman", size=20, bold=True)
        ws_bg["A5"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg["A6"] = "Kính gửi:"
        ws_bg["A7"] = "Người nhận:"
        ws_bg["A8"] = "Email/Sdt:"
        for cell_id in ["A6", "A7", "A8"]:
            ws_bg[cell_id].font = Font(name="Times New Roman", size=10, bold=True)
            ws_bg[cell_id].alignment = Alignment(horizontal="left", vertical="center")

        ws_bg["G6"] = "Người gửi:"
        ws_bg["G7"] = "Điện thoại:"
        ws_bg["G8"] = "TPHCM, ngày tháng năm 2026"
        for cell_id2 in ["G6", "G7", "G8"]:
            ws_bg[cell_id2].font = Font(name="Times New Roman", size=10, bold=True)
            ws_bg[cell_id2].alignment = Alignment(horizontal="right", vertical="center")

        ws_bg.merge_cells("A9:G9")
        ws_bg["A9"] = "Nội dung:"
        ws_bg["A9"].font = Font(name="Times New Roman", size=10, bold=True)

        thin_side = Side(style="thin", color="000000")
        for r in range(6, 9):
            for c in range(1, 8):
                cell = ws_bg.cell(row=r, column=c)
                cell.border = Border(
                    top=thin_side if r == 6 else cell.border.top,
                    bottom=thin_side if r == 8 else cell.border.bottom,
                    left=thin_side if c == 1 else cell.border.left,
                    right=thin_side if c == 7 else cell.border.right,
                )
        for c in range(1, 8):
            ws_bg.cell(row=9, column=c).border = Border(bottom=thin_side)

        ws_bg.merge_cells("A10:G10")
        ws_bg["A10"] = (
            "Cảm ơn Quý khách hàng đã quan tâm và tin tưởng sản phẩm và dịch vụ"
            " của công ty chúng tôi. Chúng tôi hân hạnh gửi đến Quý khách hàng"
            " bảng chào giá như sau:"
        )
        ws_bg["A10"].font = Font(name="Times New Roman", size=10, italic=True)
        ws_bg["A10"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws_bg.row_dimensions[10].height = 35

        headers_bg = [
            ("A11", "Stt"),
            ("B11", "Nội dung báo giá"),
            ("C11", "ĐVT"),
            ("D11", "Số lượng"),
            ("E11", "Đơn giá\n(VNĐ)"),
            ("F11", "Thuế GTGT"),
            ("G11", "Thành tiền\n(VNĐ)"),
        ]

        for cell_id, text in headers_bg:
            c = ws_bg[cell_id]
            c.value = text
            c.font = Font(name="Times New Roman", size=10, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws_bg["A12"] = 1
        ws_bg["A12"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg["B12"] = "Hệ thống\n(Xem bảng giá chi tiết đính kèm)"
        ws_bg["B12"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws_bg["C12"] = "Hệ thống"
        ws_bg["C12"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg["D12"] = 1
        ws_bg["D12"].alignment = Alignment(horizontal="center", vertical="center")

        ws_bg["E12"] = f"='GUI_KH'!I{tot_row_ct}"
        ws_bg["E12"].number_format = num_format_vnd
        ws_bg["E12"].alignment = Alignment(horizontal="right", vertical="center")

        ws_bg["F12"] = "=E12*8%"
        ws_bg["F12"].number_format = num_format_vnd
        ws_bg["F12"].alignment = Alignment(horizontal="right", vertical="center")

        ws_bg["G12"] = "=E12+F12"
        ws_bg["G12"].number_format = num_format_vnd
        ws_bg["G12"].alignment = Alignment(horizontal="right", vertical="center")

        ws_bg.row_dimensions[12].height = 35

        for r in range(11, 13):
            for col_idx in range(1, 8):
                cell = ws_bg.cell(row=r, column=col_idx)
                cell.border = thin_border
                if cell.font:
                    cell.font = Font(
                        name="Times New Roman",
                        size=10,
                        bold=cell.font.bold,
                        italic=cell.font.italic
                    )

        terms_bg = [
            ("A13:G13", "Ghi chú: Thuế GTGT tạm tính, được điều chỉnh theo quy định tại thời điểm xuất hóa đơn.", False, True),
            ("A14:G14", "Điều kiện thương mại:", True, False),
            ("A15:G15", "1. Địa điểm thực hiện:", True, False),
            ("A16:G16", "   - Phạm vi Thành phố Hồ Chí Minh", False, False),
            ("A17:G17", "2. Giá đã bao gồm:", True, False),
            ("A18:G18", "   - Chi phí vận chuyển, lắp đặt do bên Bán chịu.", False, False),
            ("A19:G19", "3. Thanh toán:", True, False),
            ("A20:G20", "   - Đợt 1: Tạm ứng 50% giá trị hợp đồng sau khi hợp đồng được ký kết.", False, False),
            ("A21:G21", "   - Đợt 2: Thanh toán 50% giá trị còn lại của hợp đồng trong vòng 15 ngày làm việc sau khi hai Bên ký Biên bản nghiệm thu hoàn thành và đưa vào sử dụng, đồng thời bên Mua đã nhận đầy đủ hồ sơ thanh toán hợp lệ của bên Bán.", False, False),
            ("A22:G22", "4. Thời gian thực hiện hợp đồng:", True, False),
            ("A23:G23", "   - Trong vòng 15 ngày kể từ ngày kí hợp đồng và bên Bán nhận được tạm ứng của bên Mua.", False, False),
            ("A24:G24", "5. Thời gian bảo hành:", True, False),
            ("A25:G25", "   - BH lắp đặt hệ thống: 12 tháng kể từ ngày nghiệm thu, bàn giao.", False, False),
            ("A26:G26", "   - BH thiết bị theo chính sách của hãng sản xuất (xem bảng giá chi tiết).", False, False),
            ("A27:G27", "6. Thời hạn chào giá:", True, False),
            ("A29:G29", "Chúng tôi rất mong nhận được sự hợp tác với Quý khách hàng!", False, True),
        ]

        for range_str, text, is_bold, is_italic in terms_bg:
            ws_bg.merge_cells(range_str)
            first_cell = ws_bg[range_str.split(":")[0]]
            first_cell.value = text
            first_cell.font = Font(
                name="Times New Roman", size=10, bold=is_bold, italic=is_italic,
                underline="single" if text == "Điều kiện thương mại:" else None
            )
            first_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws_bg.row_dimensions[21].height = 35

        ws_bg["A28"] = "   - 30 ngày"
        ws_bg["A28"].font = Font(name="Times New Roman", size=10, bold=False)
        ws_bg["A28"].alignment = Alignment(horizontal="left", vertical="center")

        ws_bg["F31"] = "Công ty TNHH Công Nghệ DVC"
        ws_bg["F31"].font = Font(name="Times New Roman", size=10, bold=True)
        ws_bg["F31"].alignment = Alignment(horizontal="center", vertical="center")

        col_widths_bg = {"A": 5, "B": 28, "C": 9, "D": 9, "E": 10, "F": 10, "G": 11}
        for col_letter, width in col_widths_bg.items():
            ws_bg.column_dimensions[col_letter].width = width

        sheet_order = ["BÁO GIÁ", "GUI_KH", "CHI TIẾT"]
        writer.book._sheets = [writer.book[s] for s in sheet_order if s in writer.book.sheetnames]

    return output.getvalue()


# --- 6.1. HÀM TẠO FILE EXCEL DÙNG CHO NHẬP TRỰC TIẾP (BAOGIA_KH - 11 CỘT) ---
def generate_direct_input_excel(raw_input_df):
    output = io.BytesIO()
    wb = Workbook()
    ws_bg = wb.active
    ws_bg.title = "BAOGIA_KH"

    # 1. THIẾT LẬP TRANG IN & KHỔ GIẤY (In từ A -> K)
    ws_bg.page_setup.orientation = ws_bg.ORIENTATION_PORTRAIT
    ws_bg.page_setup.paperSize = ws_bg.PAPERSIZE_A4
    ws_bg.sheet_properties.pageSetUpPr.fitToPage = True
    ws_bg.page_setup.fitToWidth = 1
    ws_bg.page_setup.fitToHeight = 0

    num_format_vnd = "#,##0"
    num_format_percent = "0%"
    thin_side = Side(style="thin", color="000000")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # 2. HEADER CÔNG TY & TIÊU ĐỀ (Gộp từ A -> K)
    try:
        img = Image("logo_dvc.png")
        img.width = 90
        img.height = 90
        ws_bg.add_image(img, "A1")
    except Exception:
        pass

    ws_bg["A1"] = "CÔNG TY TNHH CÔNG NGHỆ DVC"
    ws_bg["A1"].font = Font(name="Times New Roman", size=12, bold=True)
    ws_bg["A1"].alignment = align_center
    ws_bg.merge_cells("A1:K1")

    ws_bg["A2"] = "**********"
    ws_bg["A2"].alignment = align_center
    ws_bg.merge_cells("A2:K2")

    ws_bg["A3"] = "Hotline: 0909 661 579"
    ws_bg["A3"].font = Font(name="Times New Roman", size=10)
    ws_bg["A3"].alignment = align_center
    ws_bg.merge_cells("A3:K3")

    ws_bg["A4"] = "Email: dvc@dvctech.vn - Website: dvctech.vn"
    ws_bg["A4"].font = Font(name="Times New Roman", size=10, underline="single")
    ws_bg["A4"].alignment = align_center
    ws_bg.merge_cells("A4:K4")

    ws_bg["A5"] = "BẢNG BÁO GIÁ"
    ws_bg["A5"].font = Font(name="Times New Roman", size=20, bold=True)
    ws_bg["A5"].alignment = align_center
    ws_bg.merge_cells("A5:K5")

    # 3. THÔNG TIN KHÁCH HÀNG & GIAO DỊCH (Góc phải ở cột K)
    ws_bg["A6"] = "Kính gửi:"
    ws_bg["A7"] = "Người nhận:"
    ws_bg["A8"] = "Email/Sđt:"
    for cell_id in ["A6", "A7", "A8"]:
        ws_bg[cell_id].font = Font(name="Times New Roman", size=10, bold=True)
        ws_bg[cell_id].alignment = Alignment(
            horizontal="left", vertical="center"
        )

    ws_bg["K6"] = "Người gửi:"
    ws_bg["K7"] = "Điện thoại:"
    ws_bg["K8"] = "TPHCM, ngày tháng năm 2026"
    for cell_id2 in ["K6", "K7", "K8"]:
        ws_bg[cell_id2].font = Font(name="Times New Roman", size=10, bold=True)
        ws_bg[cell_id2].alignment = Alignment(
            horizontal="right", vertical="center"
        )

    ws_bg["A9"] = "Nội dung:"
    ws_bg["A9"].font = Font(name="Times New Roman", size=10, bold=True)
    ws_bg.merge_cells("A9:K9")

    # Khung viền phần thông tin khách hàng (từ cột 1 -> 11)
    for r in range(6, 9):
        for c in range(1, 12):
            cell = ws_bg.cell(row=r, column=c)
            cell.border = Border(
                top=thin_side if r == 6 else cell.border.top,
                bottom=thin_side if r == 8 else cell.border.bottom,
                left=thin_side if c == 1 else cell.border.left,
                right=thin_side if c == 11 else cell.border.right,
            )
    for c in range(1, 12):
        ws_bg.cell(row=9, column=c).border = Border(bottom=thin_side)

    ws_bg["A10"] = (
        "Cảm ơn Quý Đơn vị đã quan tâm và tin tưởng sản phẩm và dịch vụ"
        " của công ty chúng tôi. Chúng tôi hân hạnh gửi đến Quý Đơn vị bảng chào"
        " giá như sau:"
    )
    ws_bg["A10"].font = Font(name="Times New Roman", size=10, italic=True)
    ws_bg["A10"].alignment = align_left
    ws_bg.row_dimensions[10].height = 35
    ws_bg.merge_cells("A10:K10")

    # 4. TIÊU ĐỀ BẢNG DỮ LIỆU (Đúng 11 cột yêu cầu)
    headers = [
        ("A11", "STT"),
        ("B11", "Tên hàng hóa/Dịch vụ"),
        ("C11", "Mã hàng"),
        ("D11", "Hãng/\nXuất xứ"),
        ("E11", "ĐVT"),
        ("F11", "Số lượng"),
        ("G11", "Đơn giá\n(VNĐ)"),
        ("H11", "Thành tiền\n(VNĐ)"),
        ("I11", "MARGIN"),
        ("J11", "ĐG COST"),
        ("K11", "TT COST"),
    ]

    for cell_id, text in headers:
        c = ws_bg[cell_id]
        c.value = text
        c.font = Font(name="Times New Roman", size=10, bold=True)
        c.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        c.border = thin_border

    # Lọc bỏ dòng trống
    if "Thiết bị" in raw_input_df.columns:
        valid_df = raw_input_df[
            raw_input_df["Thiết bị"]
            .dropna()
            .astype(str)
            .str.strip()
            .ne("")
        ].reset_index(drop=True)
    elif "Tên hàng hóa/Dịch vụ" in raw_input_df.columns:
        valid_df = raw_input_df[
            raw_input_df["Tên hàng hóa/Dịch vụ"]
            .dropna()
            .astype(str)
            .str.strip()
            .ne("")
        ].reset_index(drop=True)
    else:
        valid_df = raw_input_df.copy()

    start_row = 12
    n_rows = len(valid_df)

    for i in range(n_rows):
        r = start_row + i
        row = valid_df.iloc[i]

        name = (
            row.get("Thiết bị")
            or row.get("Tên hàng hóa/Dịch vụ")
            or row.get("Mô tả")
            or ""
        )
        ma_hang = row.get("Mã hàng") or ""
        brand = (
            row.get("Hãng/Xuất xứ")
            or row.get("Hãng/\nXuất xứ")
            or row.get("Nhãn hiệu/Xuất xứ")
            or ""
        )
        unit = row.get("ĐVT") or "Cái"
        qty = clean_currency(row.get("Số lượng", 1))

        margin = parse_margin(
            row.get("Margin") or row.get("Margin Thiết bị") or 0
        )
        cost = parse_formula_or_value(
            row.get("Đơn giá COST") or row.get("ĐG COST Thiết bị") or row.get("Giá Cost") or 0
        )

        # Ghi các giá trị vào các cột A -> K
        ws_bg.cell(row=r, column=1, value=i + 1).alignment = align_center
        ws_bg.cell(row=r, column=2, value=name).alignment = align_left
        ws_bg.cell(row=r, column=3, value=ma_hang).alignment = align_center
        ws_bg.cell(row=r, column=4, value=brand).alignment = align_center
        ws_bg.cell(row=r, column=5, value=unit).alignment = align_center
        ws_bg.cell(row=r, column=6, value=qty).alignment = align_center

        # Đơn giá (VNĐ) = ROUNDUP(Cost / (1 - Margin), -3)
        ws_bg.cell(
            row=r, column=7, value=f"=ROUNDUP(J{r}/(1-I{r}),-3)"
        ).number_format = num_format_vnd

        # Thành tiền (VNĐ) = Số lượng * Đơn giá
        ws_bg.cell(
            row=r, column=8, value=f"=F{r}*G{r}"
        ).number_format = num_format_vnd

        # MARGIN
        ws_bg.cell(row=r, column=9, value=margin).number_format = (
            num_format_percent
        )

        # ĐG COST
        ws_bg.cell(row=r, column=10, value=cost).number_format = (
            num_format_vnd
        )

        # TT COST = Số lượng * ĐG COST
        ws_bg.cell(
            row=r, column=11, value=f"=F{r}*J{r}"
        ).number_format = num_format_vnd

        for col_idx in range(1, 12):
            c = ws_bg.cell(row=r, column=col_idx)
            c.border = thin_border
            c.font = Font(name="Times New Roman", size=10)
            if col_idx in [7, 8, 10, 11]:
                c.alignment = align_right
            elif col_idx in [1, 3, 4, 5, 6, 9]:
                c.alignment = align_center
            else:
                c.alignment = align_left

    end_device_row = start_row + n_rows - 1 if n_rows > 0 else start_row

    r_subtotal = end_device_row + 1
    r_vat = r_subtotal + 1
    r_total = r_vat + 1

    # Dòng 1: THÀNH TIỀN TRƯỚC THUẾ (Gộp A -> G, Giá trị ở cột H)
    ws_bg.cell(
        row=r_subtotal, column=1, value="THÀNH TIỀN TRƯỚC THUẾ"
    ).font = Font(name="Times New Roman", size=10, bold=True)
    ws_bg.cell(row=r_subtotal, column=1).alignment = align_center
    ws_bg.merge_cells(
        start_row=r_subtotal, start_column=1, end_row=r_subtotal, end_column=7
    )

    val_sub = ws_bg.cell(
        row=r_subtotal,
        column=8,
        value=(
            f"=SUM(H{start_row}:H{end_device_row})" if n_rows > 0 else 0
        ),
    )
    val_sub.font = Font(name="Times New Roman", size=10, bold=True)
    val_sub.alignment = align_right
    val_sub.number_format = num_format_vnd

    # Dòng 2: THUẾ GTGT
    ws_bg.cell(row=r_vat, column=1, value="THUẾ GTGT").font = Font(
        name="Times New Roman", size=10, bold=True
    )
    ws_bg.cell(row=r_vat, column=1).alignment = align_center
    ws_bg.merge_cells(
        start_row=r_vat, start_column=1, end_row=r_vat, end_column=7
    )

    val_vat = ws_bg.cell(row=r_vat, column=8, value=f"=H{r_subtotal}*8%")
    val_vat.font = Font(name="Times New Roman", size=10, bold=True)
    val_vat.alignment = align_right
    val_vat.number_format = num_format_vnd

    # Dòng 3: TỔNG CỘNG
    ws_bg.cell(row=r_total, column=1, value="TỔNG CỘNG").font = Font(
        name="Times New Roman", size=10, bold=True
    )
    ws_bg.cell(row=r_total, column=1).alignment = align_center
    ws_bg.merge_cells(
        start_row=r_total, start_column=1, end_row=r_total, end_column=7
    )

    val_tot = ws_bg.cell(
        row=r_total, column=8, value=f"=H{r_subtotal}+H{r_vat}"
    )
    val_tot.font = Font(name="Times New Roman", size=10, bold=True)
    val_tot.alignment = align_right
    val_tot.number_format = num_format_vnd

    # Kẻ khung và tô màu xám cho phần tổng tiền (Từ Cột A -> Cột K)
    gray_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )
    for r_idx in [r_subtotal, r_vat, r_total]:
        for c_idx in range(1, 12):
            cell = ws_bg.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            cell.fill = gray_fill

    # 5. ĐIỀU KIỆN THƯƠNG MẠI (Gộp từ A -> K)
    r_terms_start = r_total + 1
    terms_bg = [
        (
            f"A{r_terms_start}:K{r_terms_start}",
            "Điều kiện thương mại:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+1}:K{r_terms_start+1}",
            "1. Địa điểm thực hiện:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+2}:K{r_terms_start+2}",
            "   - Phạm vi Thành phố Hồ Chí Minh",
            False,
            False,
        ),
        (
            f"A{r_terms_start+3}:K{r_terms_start+3}",
            "2. Giá đã bao gồm:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+4}:K{r_terms_start+4}",
            "   - Chi phí vận chuyển, lắp đặt do bên Bán chịu.",
            False,
            False,
        ),
        (
            f"A{r_terms_start+5}:K{r_terms_start+5}",
            "3. Thanh toán:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+6}:K{r_terms_start+6}",
            "   - Thanh toán 100% ngay sau khi bàn giao",
            False,
            False,
        ),
        (
            f"A{r_terms_start+7}:K{r_terms_start+7}",
            "4. Thời gian thực hiện:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+8}:K{r_terms_start+8}",
            "   - Ngay sau khi xác nhận đơn hàng.",
            False,
            False,
        ),
        (
            f"A{r_terms_start+9}:K{r_terms_start+9}",
            "5. Thời gian bảo hành:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+10}:K{r_terms_start+10}",
            "   - BH thiết bị theo chính sách của hãng.",
            False,
            False,
        ),
        (
            f"A{r_terms_start+11}:K{r_terms_start+11}",
            "6. Thời hạn chào giá:",
            True,
            False,
        ),
        (
            f"A{r_terms_start+12}:K{r_terms_start+12}",
            "   - 07 ngày kể từ ngày chào giá.",
            False,
            False,
        ),
        (
            f"A{r_terms_start+13}:K{r_terms_start+13}",
            "Chúng tôi rất mong được hợp tác với Quý Đơn vị.",
            False,
            True,
        ),
    ]

    for range_str, text, is_bold, is_italic in terms_bg:
        top_left_cell_id = range_str.split(":")[0]
        first_cell = ws_bg[top_left_cell_id]
        first_cell.value = text
        first_cell.font = Font(
            name="Times New Roman",
            size=10,
            bold=is_bold,
            italic=is_italic,
            underline="single" if text == "Điều kiện thương mại:" else None,
        )
        first_cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )
        ws_bg.merge_cells(range_str)

    # Chữ ký Công ty ở cột K
    r_sign = r_terms_start + 15
    ws_bg[f"K{r_sign}"] = "Công ty TNHH Công Nghệ DVC"
    ws_bg[f"K{r_sign}"].font = Font(name="Times New Roman", size=10, bold=True)
    ws_bg[f"K{r_sign}"].alignment = align_center

    col_widths_bg = {
        "A": 6,   # STT
        "B": 35,  # Tên hàng hóa/Dịch vụ
        "C": 15,  # Mã hàng
        "D": 15,  # Hãng/Xuất xứ
        "E": 8,   # ĐVT
        "F": 9,   # Số lượng
        "G": 15,  # Đơn giá (VNĐ)
        "H": 16,  # Thành tiền (VNĐ)
        "I": 12,  # MARGIN
        "J": 15,  # ĐG COST
        "K": 16,  # TT COST
    }
    for col_letter, width in col_widths_bg.items():
        ws_bg.column_dimensions[col_letter].width = width

    wb.save(output)
    return output.getvalue()


# --- 7. GIAO DIỆN TẢI FORM MẪU & NHẬP TRỰC TIẾP ---
if "show_manual_input" not in st.session_state:
    st.session_state.show_manual_input = False

col_template, col_manual_btn = st.columns([1, 1])

with col_template:
    st.write("TẢI FILE BÁO GIÁ MẪU ĐỂ NHẬP THEO FORM HỆ THỐNG:")
    sample_file_data = generate_sample_template()
    st.download_button(
        label="Form Báo Giá Mẫu (.xlsx)",
        data=sample_file_data,
        file_name="BG Mẫu - DVC.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col_manual_btn:
    st.write("HOẶC NHẬP DỮ LIỆU BÁO GIÁ TRỰC TIẾP TRÊN WEB:")
    if st.button("Nhập Báo Giá Trực Tiếp", use_container_width=True, type="secondary"):
        st.session_state.show_manual_input = not st.session_state.show_manual_input

# --- 8. KHUNG NHẬP DỮ LIỆU BÁO GIÁ TRỰC TIẾP ---
if st.session_state.show_manual_input:
    st.markdown("---")
    st.subheader("Bảng nhập thông tin báo giá trực tiếp")
    st.caption("💡 **Mẹo:** Ở cột **Đơn giá COST**, bạn có thể nhập số trực tiếp (vd: `100000`) hoặc gõ công thức tính toán (vd: `100000/1.08` hay `50000*2`). Excel sẽ giữ nguyên công thức này khi xuất file!")

    sample_manual_df = pd.DataFrame([
        {
            "STT": 1,
            "Thiết bị": "",
            "Mã hàng": "",
            "Hãng/Xuất xứ": "",
            "Mô tả": "",
            "ĐVT": "Cái",
            "Số lượng": 1,
            "Margin": "",
            "Đơn giá COST": "0",
            "NCC": ""
        }
    ])

    column_configs = {
        "STT": st.column_config.NumberColumn("STT", width="small", min_value=1, step=1),
        "Thiết bị": st.column_config.TextColumn("Thiết bị", width="large"),
        "Mã hàng": st.column_config.TextColumn("Mã hàng", width="medium"),
        "Hãng/Xuất xứ": st.column_config.TextColumn("Hãng/Xuất xứ", width="medium"),
        "Mô tả": st.column_config.TextColumn("Mô tả", width="large"),
        "ĐVT": st.column_config.SelectboxColumn("ĐVT", width="small", options=["Cái", "Bộ", "Mét", "Cuộn", "Lô", "Hệ thống", "Gói"], required=True),
        "Số lượng": st.column_config.NumberColumn("Số lượng", width="small", min_value=1, step=1, default=1),
        "Margin": st.column_config.TextColumn("Margin", width="small"),
        "Đơn giá COST": st.column_config.TextColumn("Đơn giá COST", width="medium", help="Nhập số hoặc phép tính (vd: 100000/1.08)"),
        "NCC": st.column_config.TextColumn("NCC", width="medium"),
    }

    edited_manual_df = st.data_editor(
        sample_manual_df,
        column_config=column_configs,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key="manual_entry_editor"
    )

    if st.button("Upload Dữ Liệu", type="primary"):
        if edited_manual_df.empty:
            st.error("Vui lòng nhập dữ liệu trước khi Upload!")
        else:
            try:
                excel_bytes = generate_direct_input_excel(edited_manual_df)

                st.subheader("Xem trước dữ liệu vừa nhập:")
                st.dataframe(edited_manual_df, use_container_width=True, hide_index=True)

                st.download_button(
                    label="Xuất FILE Bảng Giá Chi Tiết",
                    data=excel_bytes,
                    file_name="BG - DVC - R1.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                )
            except Exception as e:
                st.error(f"Có lỗi khi xử lý dữ liệu nhập tay: {e}")

st.divider()

# --- 9. GIAO DIỆN UPLOAD FILE EXCEL CÓ SẴN ---
st.write("HOẶC UPLOAD FILE EXCEL CÓ SẴN TẠI ĐÂY (.xlsx, .xls):")

uploaded_file = st.file_uploader(
    "Upload", type=["xlsx", "xls"], label_visibility="collapsed"
)

if uploaded_file is not None:
    try:
        file_name = uploaded_file.name.lower()
        if file_name.endswith(".xlsx"):
            input_df = pd.read_excel(uploaded_file, engine="openpyxl")
        elif file_name.endswith(".xls"):
            try:
                input_df = pd.read_excel(uploaded_file, engine="xlrd")
            except Exception:
                uploaded_file.seek(0)
                input_df = pd.read_excel(uploaded_file, engine="openpyxl")
        else:
            input_df = pd.read_excel(uploaded_file)

        st.subheader("Xem trước dữ liệu vừa tải lên:")
        st.dataframe(input_df, use_container_width=True, hide_index=True)

        excel_bytes = process_dataframe_and_generate_excel(input_df)

        name_without_ext, ext = os.path.splitext(uploaded_file.name)
        output_filename = f"{name_without_ext} - R1{ext}"

        st.download_button(
            label="Xuất FILE Bảng Giá Chi Tiết",
            data=excel_bytes,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    except Exception as e:
        st.error(f"Có lỗi khi xử lý file: {e}")
